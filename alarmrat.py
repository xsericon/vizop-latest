# -*- coding: utf-8 -*-
# Module alarmrat: part of Vizop, (c) 2020 xSeriCon
# Implements alarm rationalization PHA model and associated Viewports

# library modules
# from __future__ import division # makes a/b yield exact, not truncated, result. Must be 1st import
import wx, wx.grid, wx.lib.gridmovers # basic GUI functions and grid widget
import zmq, copy, collections, sys
import xml.etree.ElementTree as ElementTree # XML handling
import openpyxl
from openpyxl.worksheet import dimensions as Dimensions
from openpyxl.styles import PatternFill, Border, Side, alignment, Protection, Font, fills

# other vizop modules required here
import text, utilities, core_classes, info, vizop_misc, projects, art, display_utilities
from display_utilities import UIWidgetItem

# constants applicable to alarm rationalization

AllAlarmKinds = []

class AlarmKindItem(object): # kinds of alarm, e.g. high, low, discrepancy
	def __init__(self, InternalName='', HumanName='', Tag='', MatchStrings=[], UserSelectable=True, Importable=True):
		# Tag (str): str to be inserted into alarm tag for this alarm kind
		# MatchStrings (list of str): strings that will be recognised for this AlarmKindItem in the import file (should be all lower case)
		# UserSelectable (bool): whether this AlarmKindItem will be offered to the user for selection
		# Importable (bool): whether this AlarmKindItem can be imported from an import file
		assert isinstance(InternalName, str)
		assert isinstance(HumanName, str)
		assert isinstance(MatchStrings, list)
		assert isinstance(UserSelectable, bool)
		assert isinstance(Importable, bool)
		self.InternalName = InternalName
		self.HumanName = HumanName
		self.Tag = Tag
		self.MatchStrings = MatchStrings
		self.UserSelectable = UserSelectable
		self.Importable = Importable
		AllAlarmKinds.append(self)

UndefinedAlarm = AlarmKindItem(InternalName='Undefined', HumanName=_('Undefined'), UserSelectable=False, Importable=False)
	# this is the default AlarmKindItem assigned to imported alarms if they don't match any other AlarmKindItem
FourHighAlarm = AlarmKindItem(InternalName='HHHH', HumanName=_('4 High'), Tag='HHHH', MatchStrings=['hhhh', 'hihihihi', 'highhighhighhigh', '4h'])
ThreeHighAlarm = AlarmKindItem(InternalName='HHH', HumanName=_('3 High'), Tag='HHH', MatchStrings=['hhh', 'hihihi', 'highhighhigh', '3h'])
HighHighAlarm = AlarmKindItem(InternalName='HH', HumanName=_('High high'), Tag='HH', MatchStrings=['hh', 'hihi', 'highhigh', '2h'])
HighAlarm = AlarmKindItem(InternalName='H', HumanName=_('High'), Tag='H', MatchStrings=['h', 'hi', 'high'])
LowAlarm = AlarmKindItem(InternalName='L', HumanName=_('Low'), Tag='L', MatchStrings=['l', 'lo', 'low'])
LowLowAlarm = AlarmKindItem(InternalName='LL', HumanName=_('Low low'), Tag='LL', MatchStrings=['ll', 'lolo', 'lowlow'])
ThreeLowAlarm = AlarmKindItem(InternalName='LLL', HumanName=_('3 Low'), Tag='LLL', MatchStrings=['lll', 'lololo', 'lowlowlow', '3l'])
FourLowAlarm = AlarmKindItem(InternalName='LLLL', HumanName=_('4 Low'), Tag='LLLL', MatchStrings=['llll', 'lolololo', 'lowlowlowlow', '4l'])
DiscrepancyAlarm = AlarmKindItem(InternalName='Discrepancy', HumanName=_('Discrepancy'), Tag='H', MatchStrings=['disc', 'discrep', 'discrepancy'])
DeviationAlarm = AlarmKindItem(InternalName='Deviation', HumanName=_('Deviation'), Tag='H', MatchStrings=['dev', 'deviation'])
LowDeltaAlarm = AlarmKindItem(InternalName='LowD', HumanName=_('Low rate of change'), Tag='L',
	MatchStrings=['deltalow', 'lowdelta', 'Δl', 'lΔ', 'low rate of change', 'rate of change low'])
HighDeltaAlarm = AlarmKindItem(InternalName='HighD', HumanName=_('High rate of change'), Tag='H', MatchStrings=['delta', 'Δ', 'rate of change'])
FaultAlarm = AlarmKindItem(InternalName='Fault', HumanName=_('Fault'), MatchStrings=['fault', 'err', 'warn'])
OtherAlarm = AlarmKindItem(InternalName='Other', HumanName=_('Other'), MatchStrings=['other', 'nd', 'na', 'n/a', '-', 'later'])

def GetAlarmKind(KindStr, ImportableOnly=False): # find AlarmKindItem whose MatchStrings contains KindStr, and return it.
	# Returns None if no match found
	# ImportableOnly (bool): if True, only attempt to match alarm kinds with Importable == True
	# Assumes KindStr is already stripped and converted to lower case
	assert isinstance(KindStr, str)
	assert isinstance(ImportableOnly, bool)
	Match = None
	for ThisKind in AllAlarmKinds:
		if KindStr in ThisKind.MatchStrings:
			if ThisKind.Importable or not ImportableOnly:
				Match = ThisKind
			break # no need to search any more, even if we found a non-importable alarm kind
	return Match

class RationalizationStatusItem(object): # status of an individual alarm, e.g. "not rationalized"
	def __init__(self, InternalName='', HumanName='', UserSelectable=True, Importable=True, **Attribs):
		# UserSelectable (bool): whether this status will be offered to the user for selection
		# Importable (bool): whether this status can be imported from an import file
		assert isinstance(InternalName, str)
		assert isinstance(HumanName, str)
		assert isinstance(UserSelectable, bool)
		assert isinstance(Importable, bool)
		self.InternalName = InternalName
		self.HumanName = HumanName # used as recognition key for import
		self.UserSelectable = UserSelectable
		self.Importable = Importable
		for a, v in Attribs.items(): setattr(self, a, v)

NotRationalizedStatus = RationalizationStatusItem(InternalName='NotRationalized', HumanName=_('Not rationalized'),
	UserSelectable=True, Importable=True)
# TODO: add other status definitions

class ClassificationItem(object): # classifications to which alarms can be assigned, e.g. "environmental"
	def __init__(self, InternalName='', HumanName='', UserSelectable=True, Importable=True, **Attribs):
		# UserSelectable (bool): whether this classification will be offered to the user for selection
		# Importable (bool): whether this classification can be imported from an import file
		assert isinstance(InternalName, str)
		assert isinstance(HumanName, str)
		assert isinstance(UserSelectable, bool)
		assert isinstance(Importable, bool)
		self.InternalName = InternalName
		self.HumanName = HumanName # used as recognition key for import
		self.UserSelectable = UserSelectable
		self.Importable = Importable
		for a, v in Attribs.items(): setattr(self, a, v)

# library of default classes
NormalClass = ClassificationItem(InternalName='Class1', HumanName=_('Default'))
IPLClass = ClassificationItem(InternalName='Class2', HumanName=_('IPL for SIF'))
SafetyCritClass = ClassificationItem(InternalName='Class3', HumanName=_('Safety critical'))
EnvClass = ClassificationItem(InternalName='Class4', HumanName=_('Environmental requirement'))
HighMaintClass = ClassificationItem(InternalName='Class5', HumanName=_('High maintenance'))
MachineProtClass = ClassificationItem(InternalName='Class6', HumanName=_('Machine protection'))
FnGClass = ClassificationItem(InternalName='Class7', HumanName=_('Fire & gas'))
SystemsClass = ClassificationItem(InternalName='Class8', HumanName=_('Systems fault'))
# if the user creates new classifications, the InternalName can be 'Class' + next available int

DefaultClassifications = [NormalClass, IPLClass, SafetyCritClass, EnvClass, HighMaintClass, MachineProtClass, FnGClass,
	SystemsClass]

def GetCategoryFromStr(Instr='', Categories=[], MatchAttribName=''):
	# find the category in Categories whose attrib (list of lowercase str) named in MatchAttribName contains Instr (str)
	# Assumes InStr is already stripped and converted to lower case
	# returns matching category instance or None if no match found
	assert isinstance(Instr, str)
	assert isinstance(Categories, list)
	assert isinstance(MatchAttribName, str)
	assert hasattr(Categories[0], MatchAttribName)
	Match = None
	for ThisCateg in Categories:
		if Instr in getattr(ThisCateg, MatchAttribName):
			Match = ThisCateg
			break  # no need to search any more
	return Match

class AlarmAttrib(object): # single attribute of an alarm. Used for making lists of attribs for import
	AllAlarmAttribs = [] # register of all attribs defined in this class

	def __init__(self, AttribName='', HumanName='', AttribType=str, ImportDirectly=True, **Attribs):
		# AttribName: the name of corresponding attrib in AlarmObjectInCore, if any
		# HumanName: name for display
		# AttribType: type of value to be imported for this attrib
		# ImportDirectly (bool): whether to import directly into corresponding Attrib of AlarmObjectInCore
		object.__init__(self)
		AlarmAttrib.AllAlarmAttribs.append(self)
		self.AttribName = AttribName
		self.HumanName = HumanName
		self.AttribType = AttribType
		self.ImportDirectly = ImportDirectly
		for a in Attribs: setattr(self, a, Attribs[a])

# Attribs are listed in the import choices in the order below
InstrTagAttrib = AlarmAttrib(AttribName='InstrTag', HumanName=_('Instrument tag'))
AlarmTagAttrib = AlarmAttrib(AttribName='AlarmTag', HumanName=_('Alarm tag'))
ProcessUnitAttrib = AlarmAttrib(AttribName='ProcessUnit', HumanName=_('ProcessUnit'))
AlarmKindAttrib = AlarmAttrib(AttribName='AlarmKind', HumanName=_('Type of alarm'), AttribType=AlarmKindItem)
SetpointAttrib = AlarmAttrib(AttribName='Setpoint', HumanName=_('Setpoint'))
SetpointHHAttrib = AlarmAttrib(AttribName='SetpointHH', HumanName=_('Setpoint HH'), ImportDirectly=False, AlarmKind=HighHighAlarm)
SetpointHAttrib = AlarmAttrib(AttribName='SetpointH', HumanName=_('Setpoint H'), ImportDirectly=False, AlarmKind=HighAlarm)
SetpointLAttrib = AlarmAttrib(AttribName='SetpointL', HumanName=_('Setpoint L'), ImportDirectly=False, AlarmKind=LowAlarm)
SetpointLLAttrib = AlarmAttrib(AttribName='SetpointLL', HumanName=_('Setpoint LL'), ImportDirectly=False, AlarmKind=LowLowAlarm)
EngUnitAttrib = AlarmAttrib(AttribName='EngUnit', HumanName=_('Engineering unit'), ImportDirectly=False)
IsDiscreteAttrib = AlarmAttrib(AttribName='IsDiscrete', HumanName=_('Discrete?'), AttribType=bool)
CommentAttrib = AlarmAttrib(AttribName='Comment', HumanName=_('Comment'))
CausesAttrib = AlarmAttrib(AttribName='Causes', HumanName=_('Causes'), AttribType=list)
ConsequencesAttrib = AlarmAttrib(AttribName='Consequences', HumanName=_('Consequences of no response'), AttribType=list)
ConfirmationAttrib = AlarmAttrib(AttribName='Confirmation', HumanName=_('Confirmation of actual alarm'))
OpActionsAttrib = AlarmAttrib(AttribName='OpActions', HumanName=_('Operator actions'), AttribType=list)
ResponseTimeRequiredInMinAttrib = AlarmAttrib(AttribName='ResponseTimeReqInMin',
	HumanName=_('Response time required (mins)'), ImportDirectly=False, Unit=core_classes.MinuteUnit)
ResponseTimeAvailableInMinAttrib = AlarmAttrib(AttribName='ResponseTimeAvailInMin',
	HumanName=_('Response time available (mins)'), ImportDirectly=False, Unit=core_classes.MinuteUnit)
ResponseTimeRequiredCategAttrib = AlarmAttrib(AttribName='ResponseTimeReqCateg', HumanName=_('Response time required (category)'), ImportDirectly=False)
ResponseTimeAvailableCategAttrib = AlarmAttrib(AttribName='ResponseTimeAvailCateg', HumanName=_('Response time available (category)'), ImportDirectly=False)
SeverityAttrib = AlarmAttrib(AttribName='Severities', HumanName=_('Severity of consequence'), ImportDirectly=False)
PriorityAttrib = AlarmAttrib(AttribName='UserPriority', HumanName=_('Overridden priority'), ImportDirectly=False)
# Classification1Attrib = AlarmAttrib(AttribName='Class1', HumanName=_('Classification 1'), ImportDirectly=False)
# Classification2Attrib = AlarmAttrib(AttribName='Class2', HumanName=_('Classification 2'), ImportDirectly=False)
# Classification3Attrib = AlarmAttrib(AttribName='Class3', HumanName=_('Classification 3'), ImportDirectly=False)
# Classification4Attrib = AlarmAttrib(AttribName='Class4', HumanName=_('Classification 4'), ImportDirectly=False)
# Classification5Attrib = AlarmAttrib(AttribName='Class5', HumanName=_('Classification 5'), ImportDirectly=False)
# Classification6Attrib = AlarmAttrib(AttribName='Class6', HumanName=_('Classification 6'), ImportDirectly=False)
# Classification7Attrib = AlarmAttrib(AttribName='Class7', HumanName=_('Classification 7'), ImportDirectly=False)
# Classification8Attrib = AlarmAttrib(AttribName='Class8', HumanName=_('Classification 8'), ImportDirectly=False)
OnDelayAttrib = AlarmAttrib(AttribName='OnDelay', HumanName=_('On delay (sec)'), AttribType=float, Unit=core_classes.SecondUnit)
OffDelayAttrib = AlarmAttrib(AttribName='OffDelay', HumanName=_('Off delay (sec)'), AttribType=float, Unit=core_classes.SecondUnit)
PVAveragingAttrib = AlarmAttrib(AttribName='PVAveraging', HumanName=_('PV averaging timebase (sec)'), AttribType=float, Unit=core_classes.SecondUnit)
DeadbandAttrib = AlarmAttrib(AttribName='Deadband', HumanName=_('Deadband'), AttribType=float, Unit=core_classes.ProbabilityUnit)
DeadbandUnitAttrib = AlarmAttrib(AttribName='DeadbandUnit', HumanName=_('Deadband units'))

MultiLevelSetpointAttribs = [SetpointHHAttrib, SetpointHAttrib, SetpointLAttrib, SetpointLLAttrib]

class AlarmListGridTable(wx.grid.GridTableBase): # object containing data table for alarm list display

	def __init__(self, Log):
		wx.grid.GridTableBase.__init__(self)
		self.Log = Log
		self.identifiers = [InstrTagAttrib.AttribName, AlarmTagAttrib.AttribName] # internal names of columns
		self.rowLabels = []
		self.colLabels = []
		self.data = []

	def GetNumberRows(self): # required method for this class
		return len(self.data)

	def GetNumberCols(self): # required method for this class
		return len(self.identifiers)

	def IsEmptyCell(self, Row, Col): # required method for this class
		assert isinstance(Row, int)
		assert isinstance(Col, int)
		assert 0 <= Row < self.GetNumberRows()
		assert 0 <= Col < self.GetNumberCols()
		return not self.data[Row][self.identifiers[Col]]

	def GetValue(self, Row, Col): # required method for this class
		assert isinstance(Row, int)
		assert isinstance(Col, int)
		assert 0 <= Row < self.GetNumberRows() + 1 # +1 needed in case user drags an alarm to the bottom of the table
		assert 0 <= Col < self.GetNumberCols()
		return self.data[Row][self.identifiers[Col]]

	def SetValue(self, Row, Col, Value): # required method for this class
		assert isinstance(Row, int)
		assert isinstance(Col, int)
		assert 0 <= Row < self.GetNumberRows()
		assert 0 <= Col < self.GetNumberCols()
		self.data[Row][self.identifiers[Col]] = Value

	def GetRowLabelValue(self, Row):
		assert isinstance(Row, int)
		assert 0 <= Row
		return self.rowLabels[Row]

	def GetColLabelValue(self, Col):
		assert isinstance(Col, int)
		assert 0 <= Col < self.GetNumberCols()
		return self.colLabels[Col]

	def MoveColumn(self, FromColNo, ToColNo):
		assert isinstance(FromColNo, int)
		assert isinstance(ToColNo, int)
		assert 0 <= FromColNo < self.GetNumberCols()
		assert 0 <= ToColNo < self.GetNumberCols()
		# move column identifier in the list
		self.identifiers.insert(ToColNo, self.identifiers.pop(FromColNo))
		# notify grid of change
		ThisGrid = self.GetView()
		ThisGrid.BeginBatch()
		ThisGrid.ProcessTableMessage(wx.grid.GridTableMessage(self, wx.grid.GRIDTABLE_NOTIFY_COLS_DELETED, FromColNo, 1))
		ThisGrid.ProcessTableMessage(wx.grid.GridTableMessage(self, wx.grid.GRIDTABLE_NOTIFY_COLS_INSERTED, ToColNo, 1))
		ThisGrid.EndBatch()

	def MoveRow(self, FromRowNo, ToRowNo):
		assert isinstance(FromRowNo, int)
		assert isinstance(ToRowNo, int)
		assert 0 <= FromRowNo < self.GetNumberRows()
		assert 0 <= ToRowNo < self.GetNumberRows() + 1 # +1 in case the row is dragged to the bottom of the table
		# notify grid of change
		ThisGrid = self.GetView()
		ThisGrid.BeginBatch()
		ThisGrid.ProcessTableMessage(wx.grid.GridTableMessage(self, wx.grid.GRIDTABLE_NOTIFY_ROWS_DELETED, FromRowNo, 1))
		ThisGrid.ProcessTableMessage(wx.grid.GridTableMessage(self, wx.grid.GRIDTABLE_NOTIFY_ROWS_INSERTED, ToRowNo, 1))
		ThisGrid.EndBatch()

class DraggableGrid(wx.grid.Grid): # wx.grid object whose rows can be reordered by dragging

	def __init__(self, Parent, Log=sys.stdout):
		wx.grid.Grid.__init__(self, Parent, -1)
		self.DataTable = AlarmListGridTable(Log) # create a data table object
		self.SetTable(self.DataTable, True, selmode=wx.grid.Grid.SelectRows) # can select rows, but not individual cells
		# enable columns to be dragged to reorder (not currently used)
		# wx.lib.gridmovers.GridColMover(self)
		# self.Bind(wx.lib.gridmovers.EVT_GRID_COL_MOVE, self.OnColMove, self)
		# enable rows to be dragged to reorder (to implement later)
#		wx.lib.gridmovers.GridRowMover(self)
#		self.Bind(wx.lib.gridmovers.EVT_GRID_ROW_MOVE, self.OnRowMove, self)
		self.DisableCellEditControl() # disallow editing of cells
		# send events to parent window (e.g. data entry panel) for processing, if parent has a handler
		# (the hasattr() check is in case we use a grid in any other window in future, and don't implement handler)
		if hasattr(Parent, 'OnGridMouseDoubleLClick'):
			self.Bind(wx.grid.EVT_GRID_CELL_LEFT_DCLICK, Parent.OnGridMouseDoubleLClick)
		if hasattr(Parent, 'OnGridRangeSelect'):
			self.Bind(wx.grid.EVT_GRID_RANGE_SELECT, Parent.OnGridRangeSelect)

	def OnColMove(self, Event): # handle dragging of table column. Not currently used
		self.GetTable().MoveColumn(Event.GetMoveColumn(), Event.GetBeforeColumn())

	def OnRowMove(self, Event): # handle dragging of table row. Nor currently used
		self.GetTable().MoveRow(Event.GetMoveRow(), Event.GetBeforeRow())
		# call row move handler in the hosting panel
		if hasattr(self.Parent, 'MoveGridRow'):
			self.Parent.MoveGridRow(FromRowIndex=Event.GetMoveRow(), ToRowIndex=Event.GetBeforeRow())


class AlarmObjectInCore(object): # defines a single alarm in an alarm list%%%
	AllAlarmObjects = [] # register of all instances defined
	DefaultAlarmKind = UndefinedAlarm
	DefaultRationalizationStatus = NotRationalizedStatus

	def __init__(self, HostingARObject, AlarmTag=''):
		assert isinstance(HostingARObject, ARObjectInCore)
		assert isinstance(AlarmTag, str)
		assert AlarmTag # ensure it's not empty
		object.__init__(self)
		self.ID = str(utilities.NextID(AlarmObjectInCore.AllAlarmObjects)) # generate unique ID; stored as str
		AlarmObjectInCore.AllAlarmObjects.append(self) # add instance to register; must do after assigning self.ID
		self.HostingARObj = HostingARObject # parent AR containing this alarm
		# if any attrib names are changed below, change also in corresponding AlarmAttrib instance
		self.ProcessUnit = '' # name of process unit in which the alarm is located (str)
		self.InstrTag = '' # tag of instrument that generates the alarm (str)
		self.AlarmTag = AlarmTag # tag of this alarm; it's the alarm's unique identifier; assume stripped (str)
		self.AlarmKind = AlarmObjectInCore.DefaultAlarmKind
		self.Setpoint = core_classes.UserNumValueItem()
		self.IsDiscrete = False
		self.Comment = '' # str
		self.RationalizationStatus = AlarmObjectInCore.DefaultRationalizationStatus # whether alarm has been rationalized
		self.Causes = [] # strings
		self.Consequences = [] # strings, one per item in Causes
		self.Confirmation = ''
		self.OperatorActions = [] # strings, one per item in Causes
		self.ResponseTimeRequired = core_classes.UserNumValueItem()
		self.ResponseTimeAvail = core_classes.UserNumValueItem()
		self.Severities = [] # severity categories
		self.UserPriority = core_classes.CategoryNameItem() # priority defined by user, as override of calculated priority
		# (Calculated priority is calculated on the fly and not stored)
		self.OverridePriority = False # whether to use UserPriority instead of calculated priority
		self.Classifications = self.MakeDefaultClassifications(Classifications=HostingARObject.Classifications)
		# Classifications is OrderedDict: keys: classification instances; values: bool (whether the alarm belongs to this classif.)
		# If a classification is missing from the dict, it's assumed the value is False
		# We use OrderedDict because the order needs to match Classif1, Classif2 etc. for import.
		self.OnDelay = core_classes.UserNumValueItem()
		self.OffDelay = core_classes.UserNumValueItem()
		self.PVAveraging = core_classes.UserNumValueItem()
		self.Deadband = core_classes.UserNumValueItem() # this stores the numerical value, always with Probability units
			# (i.e. range 0~1). The actual unit is stored as a string in self.DeadbandUnit
		self.DeadbandUnit = ''
		self.Enabled = True # whether the alarm is switched on in the PCS

	def MakeDefaultClassifications(self, Classifications): # make and return dict of default classifications
		return collections.OrderedDict([ (c, False) for c in DefaultClassifications ])

	def GetAutoPriority(self):
		# get and return the priority (category object instance) selected from the alarm's priority matrix
		# If priority cannot be determined, returns undefined priority category
		# Look up the priority for each severity/urgency pair
		Priorities = []
		for (ThisSeverity, ThisUrgency) in zip(self.Severities, self.ResponseTimeAvail):
			# find category matching response time provided
			UrgencyCategory = vizop_misc.GetCategoryFromValue(self.HostingARObj.Urgencies, ThisUrgency, RoundUp=True)
			if UrgencyCategory: # matching urgency category found? Look up priority from severity and urgency
				Priorities.append(self.HostingARObj.PriorityMatrix.Lookup([ThisSeverity, UrgencyCategory] ))
		if Priorities: # any priorities obtained?
			return vizop_misc.HighestCategoryAmong(self.HostingARObj.Priorities,
				self.HostingARObj.PrioritiesAreInAscendingOrder, Priorities)
		else: return vizop_misc.UndefinedCategoryAmong(self.HostingARObj.Priorities)

def CreateNewAlarm(HostingARObject=None, AlarmTag=''): # create new AlarmObjectInCore() and return it
	return AlarmObjectInCore(HostingARObject=HostingARObject, AlarmTag=AlarmTag)

class ARObjectInCore(core_classes.PHAModelBaseClass):
	# defines alarm rationalization object as stored in DataCore. This represents an entire AR dataset.
	# It's handled and accessed only by DataCore, not by Viewports.
	IsBaseClass = False # must do this for every PHAModelBaseClass subclass
	PreferredKbdShortcut = 'A'
	HumanName = _('Alarm rationalization')
	InternalName = 'AR'
	AllARObjects = [] # register of all instances defined
#	ARObjectInCore.DefaultViewportType = AlarmList # now set in AlarmList declaration; can't do it here, as AlarmList is not declared yet

	def __init__(self, Proj, **Args):
		core_classes.PHAModelBaseClass.__init__(self, Proj, **Args)
		# ID is already assigned in PHAModelBaseClass.__init__
		# self.EditAllowed attrib is inherited from base class
		ARObjectInCore.AllARObjects.append(self) # add self to register; must do after assigning self.ID
		self.MaxElementID = 0 # Highest ID (int) of all elements in this AR. Used only to determine next available ID
		# define object-wide attributes
		self.ARName = ''
		self.Rev = ''
		self.BackgColour = '0,0,0' # string of rgb values (no point in using wx.Colour or tuple as we have to convert it)
		self.TextColour = '255,255,255'
		# data content of ARObjectInCore
		self.Alarms = [] # contains AlarmObjectInCore instances
		# define a list of default priorities
		self.Priorities = [
			core_classes.CategoryNameItem(XMLName='P0', HumanName=_('Critical'), HumanDescription=_('Must address immediately')),
			core_classes.CategoryNameItem(XMLName='P1', HumanName=_('High'), HumanDescription=_('Must address promptly')),
			core_classes.CategoryNameItem(XMLName='P2', HumanName=_('Medium'), HumanDescription=_('Should address soon')),
			core_classes.CategoryNameItem(XMLName='P3', HumanName=_('Low'), HumanDescription=_('Should address when available')),
			core_classes.CategoryNameItem(XMLName='P4', HumanName=_('Journal'), HumanDescription=_('Journal only')),
			core_classes.CategoryNameItem(XMLName='PU', HumanName=_('<Undefined>'), HumanDescription=_('Undefined'), IsUndefined=True)
			]
		self.PrioritiesAreInAscendingOrder = False # whether priorities in self.Priorities are in ascending order
		self.Classifications = DefaultClassifications
		# Make priority matrix (LookupTableItem instance)
		self.PriorityMatrix, self.Urgencies, self.Severities = self.MakeDefaultPriorityMatrix(Priorities=self.Priorities)
		# names of attribs referring to header, whose value is the same as the text in the XML tag
		self.TextComponentNames = ['ARName', 'Rev', 'BackgColour', 'TextColour']

	def MakeDefaultPriorityMatrix(self, Priorities=[]):
		# make a default priority lookup matrix. Return the matrix object (LookupTableItem instance), and lists of
		# urgencies and severities (CategoryNameItem instances)
		# Priorities (list of CategoryNameItem instances): priorities to use
		MyMatrix = core_classes.LookupTableItem()
		MyMatrix.HowManyDimensions = 2
		MyMatrix.DimensionHumanNames = [_('Urgency'), _('Severity')]
		MyMatrix.SeverityDimensionIndex = 1
		MyMatrix.DimensionUnits = [core_classes.DimensionlessUnit, core_classes.DimensionlessUnit]
		Urgency0 = core_classes.CategoryNameItem(XMLName='Urgent', HumanName=_('Urgent'), HumanDescription=_('<5 mins'))
		Urgency1 = core_classes.CategoryNameItem(XMLName='Soon', HumanName=_('Soon'), HumanDescription=_('5~15 mins'))
		Urgency2 = core_classes.CategoryNameItem(XMLName='Later', HumanName=_('Later'), HumanDescription=_('>15 mins'))
		UrgencyU = core_classes.CategoryNameItem(XMLName='Undefined', HumanName=_('Undefined'),
			HumanDescription=_('Undefined'), IsUndefined=True)
		Urgencies = [Urgency0, Urgency1, Urgency2, UrgencyU]
		# set numerical limits on the response time bands of urgency categories
		for ThisUrgency, MinResponseTime, MaxResponseTime in [(Urgency0, 0, 5), (Urgency1, 5, 15), (Urgency2, 15, 999)]:
			ThisUrgency.MinValue.SetMyValue(MinResponseTime)
			ThisUrgency.MinValue.SetMyUnit(core_classes.MinuteUnit)
			ThisUrgency.MaxValue.SetMyValue(MaxResponseTime)
			ThisUrgency.MaxValue.SetMyUnit(core_classes.MinuteUnit)
		Severity0 = core_classes.CategoryNameItem(XMLName='Negligible', HumanName=_('Negligible'),
												  HumanDescription=_('No significant impact'))
		Severity1 = core_classes.CategoryNameItem(XMLName='Minor', HumanName=_('Minor'),
												  HumanDescription=_('Small, reversible impact'))
		Severity2 = core_classes.CategoryNameItem(XMLName='Moderate', HumanName=_('Moderate'),
												  HumanDescription=_('Significant impact'))
		Severity3 = core_classes.CategoryNameItem(XMLName='Severe', HumanName=_('Severe'),
												  HumanDescription=_('Major impact with long-term consequences'))
		SeverityU = core_classes.CategoryNameItem(XMLName='Undefined', HumanName=_('Undefined'),
			HumanDescription=_('Undefined'), IsUndefined=True)
		Severities = [Severity0, Severity1, Severity2, Severity3, SeverityU]
		MyMatrix.Keys = [Urgencies, Severities]
		# populate priority values into matrix
		PrioUrgency0 = [Priorities[4], Priorities[1], Priorities[0], Priorities[0]]
		PrioUrgency1 = [Priorities[4], Priorities[2], Priorities[1], Priorities[0]]
		PrioUrgency2 = [Priorities[4], Priorities[3], Priorities[2], Priorities[1]]
		MyMatrix.Values = [PrioUrgency0, PrioUrgency1, PrioUrgency2]
		MyMatrix.DefaultValue = Priorities[0] # return Critical if correct value can't be read; conservative
		return MyMatrix, Urgencies, Severities

	def GetFullRedrawData(self, Viewport=None, ViewportClass=None, **Args):
		# return all data in ARObjectInCore as an XML tree, for sending to Viewport to fully render the alarm view
		# Viewport (instance of ViewportShadow): the Viewport to be displayed
		# ViewportClass (subclass of ViewportBaseClass): the class of the displayable Viewport
		# Args: any additional data to determine what data to return; depends on class of Viewport

		def PopulateOverallData(Root=None):
			# write data relating to alarm list as a whole into XML element Root
			# add alarm list object's ID as text of Root
			Root.text = self.ID
			# add alarm counts
			AlarmCountEl = ElementTree.SubElement(Root, info.AlarmCountTag)
			TotalAlarmCountEl = ElementTree.SubElement(AlarmCountEl, info.TotalAlarmCountTag)
			TotalAlarmCountEl.text = len(self.Alarms)
			# add category names to offer in choices
			for ThisTag, ThisCategoryList in [ (info.SeverityCategoriesTag, self.Severities),
					(info.PrioritiesTag, self.Priorities), (info.ClassificationsTag, self.Classifications),
					(info.ResponseTimeCategoriesTag, self.Urgencies) ]:
				for ThisCategory in ThisCategoryList:
					CategoryEl = ElementTree.SubElement(Root, ThisTag)
					CategoryEl.text = ThisCategory.InternalName
					NameEl = ElementTree.SubElement(CategoryEl, info.NameTag)
					NameEl.text = ThisCategory.HumanName
					DescriptionEl = ElementTree.SubElement(CategoryEl, info.DescriptionTag)
					DescriptionEl.text = ThisCategory.HumanName

		def PopulateAlarmData(Root, ThisAlarm):
			# write data for one alarm, ThisAlarm, into a new element in Root
			AlarmEl = ElementTree.SubElement(Root, info.AlarmTag)
			# write tags for each text attribute
			for ThisTag, ThisAttribName in [ (info.AlarmTagTag, 'AlarmTag'), (info.InstrTagTag, 'InstrTag'),
				(info.ProcessUnitTag, 'ProcessUnit'), (info.CommentTag, 'Comment'),
				(info.ConfirmationTag, 'Confirmation') ]:
				AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
				AttribEl.text = getattr(ThisAlarm, ThisAttribName)
			# write tags for each HumanName attribute
			for ThisTag, ThisAttribName in [ (info.AlarmKindTag, 'AlarmKind'),
				(info.RationalizationStatusTag, 'RationalizationStatus'),
				(info.UserPriorityTag, 'UserPriority')
				]:
				AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
				AttribEl.text = getattr(ThisAlarm, ThisAttribName).HumanName
			# write tags for each numerical attribute
			for ThisTag, ThisAttribName in [ (info.SetpointTag, 'Setpoint'), (info.OnDelayTag, 'OnDelay'),
											 (info.OffDelayTag, 'OffDelay'), (info.PVAveragingTag, 'PVAveraging'),
											 (info.DeadbandTag, 'Deadband') ]:
				AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
				AttribEl.text = getattr(ThisAlarm, ThisAttribName).GetDisplayValue()
			# write tags for each unit attribute
			for ThisTag, ThisAttribName in [ (info.SetpointUnitTag, 'Setpoint'), (info.DeadbandUnitTag, 'Deadband')
				]:
				AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
				AttribEl.text = getattr(ThisAlarm, ThisAttribName).GetMyUnit().HumanName
			# write tags for each boolean attribute
			for ThisTag, ThisAttribName in [ (info.IsDiscreteTag, 'IsDiscrete'), (info.EnabledTag, 'Enabled') ]:
				AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
				AttribEl.text = utilities.Bool2Str(getattr(ThisAlarm, ThisAttribName), TrueStr=info.TrueInXML,
					FalseStr=info.FalseInXML)
			# write tags with multiple texts in a list
			for ThisTag, ThisAttribName in [ (info.CauseTag, 'Causes'), (info.ConsequenceTag, 'Consequences'),
				(info.OperatorActionTag, 'OperatorActions'), (info.SeverityTag, 'Severities') ]:
				for ThisInstance in getattr(ThisAlarm, ThisAttribName):
					AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
					AttribEl.text = ThisInstance
			# write tags with multiple numerical values in a list
			for ThisTag, ThisAttribName in [ (info.ResponseTimeRequiredTag, 'ResponseTimeRequired'),
											 (info.ResponseTimeAvailableTag, 'ResponseTimeAvailable') ]:
				for ThisInstance in getattr(ThisAlarm, ThisAttribName):
					#J: Indentation fix
					AttribEl = ElementTree.SubElement(AlarmEl, ThisTag)
					AttribEl.text = ThisInstance.GetDisplayValue() # TODO display units?
			# write tags for each classification the alarm belongs to
			for (ThisClassification, Belongs) in self.Classifications.items():
				if Belongs: # is the alarm in this classification?
					AttribEl = ElementTree.SubElement(AlarmEl, info.ClassificationTag)
					AttribEl.text = ThisClassification.HumanName
			# write auto (looked-up) priority
			AttribEl = ElementTree.SubElement(AlarmEl, info.AutoPriorityTag)
			AttribEl.text = ThisAlarm.GetAutoPriority().HumanName
			# write suppression text (placeholder for later when we define it properly)
			AttribEl = ElementTree.SubElement(AlarmEl, info.SuppressionTag)
			AttribEl.text = 'Coming soon'
			return AlarmEl

		# GetFullRedrawData main procedure
		print("AR409 in GetFullRedrawData, coding in progress")
		# %%% finished coding here 18 Feb. Next: go to the place that calls this, and unpack the received data
		# First, make the root element
		RootElement = ElementTree.Element(ViewportClass.InternalName)
		# populate with overall alarm list data
		PopulateOverallData(RootElement)
		# populate with data for each alarm
		for ThisAlarm in self.Alarm:
			AlarmEl = PopulateAlarmData(Root=RootElement, ThisAlarm=ThisAlarm)
		return RootElement

	def DoImport(self, Pathname='', WorksheetName='', Overwrite=False, HeaderRowCount=0, MultiLevel=False,
		InstrumentTagTemplate='', AlarmTagTemplate='', ColAttribNames=[]):
		# execute import of alarms from PathName into alarm list.
		# PathName (str): full path of a readable .xlsx file
		# WorksheetName (str): name of the worksheet in the file to read
		# Overwrite (bool): whether to overwrite data in existing alarms in the alarm list, if imported alarms have the
		#	same AlarmTag
		# HeaderRowCount (int): how many header rows to skip in the file, starting from the first nonblank row
		# MultiLevel (bool): whether to read multiple alarms from one row (e.g. HH, H, L, LL)
		# InstrumentTagTemplate (str) and AlarmTagTemplate (str): template tags to deduce alarm tag from, if no column
		#	is labelled as containing AlarmTag
		# ColAttribNames (list of str): attribs to import the columns into, starting from the first nonblank column;
		#	columns to be ignored should have '' entry in this list
		# TODO undo; all data changes are captured in UndoInfo, but we're not yet putting it in an undo record
		# return dict: 'ProblemReport': (str) any problem message, human-readable (other keys to add)

		def GetStringsToDeduceAlarmTag(InstrumentTagTemplate='', AlarmTagTemplate='', AlarmKindMarker='H'):
			# get substrings needed to deduce alarm tag from instrument tag. See algorithm in spec 3-10
			# Returns: Success (bool)
			# 	prefix end marker, residue prefix, residue suffix, suffix start marker (all str)
			Success = True # whether we can deduce required substrings from inputs
			PrefixEndMarker = ResiduePrefix = ResidueSuffix = SuffixStartMarker = ''
			InstrTagStripped = utilities.StripSpaces(InstrumentTagTemplate, CharsToStrip=' ' + chr(27))
			AlarmTagStripped = utilities.StripSpaces(AlarmTagTemplate, CharsToStrip=' ' + chr(27))
			Success = bool(InstrTagStripped) and bool(AlarmTagStripped) # no empty sample tags supplied
			if Success:
				# if InstrTagStripped is entirely contained in AlarmTagStripped, set a dummy prefix end marker
				if InstrTagStripped in AlarmTagStripped:
					Prefix = InstrTagStripped
					PrefixEndMarker = chr(1)
				else:
					# find prefix end marker - the first char after the longest string, starting from left of
					# InstrTagStripped, that's present in AlarmTagStripped
					PrefixLength = 0
					Matched = True
					while PrefixLength < len(InstrTagStripped) and Matched:
						PrefixLength += 1
						Matched = InstrTagStripped[:PrefixLength] in AlarmTagStripped
					Prefix = InstrTagStripped[:PrefixLength]
					PrefixEndMarker = InstrTagStripped[PrefixLength]
				# find suffix - longest string, starting from the right of InstrTagStripped, that's present in AlarmTagStripped
				# there's no suffix if the prefix already contains the whole of AlarmTagStripped
				if (AlarmTagStripped in Prefix) or (InstrTagStripped[:-1] not in AlarmTagStripped):
					Suffix = SuffixStartMarker = ''
				else:
					SuffixLength = -1 # count backwards from the end of InstrTagStripped
					Matched = True
					while (SuffixLength > PrefixLength - len(InstrTagStripped)) and Matched:
						Matched = InstrTagStripped[SuffixLength:] in AlarmTagStripped
						SuffixLength -= 1
					Suffix = InstrTagStripped[SuffixLength+1:]
					SuffixStartMarker = InstrTagStripped[SuffixLength+1]
				# find residue - the part of AlarmTagTemplate between Prefix and Suffix
				Residue = AlarmTagStripped[AlarmTagStripped.index(Prefix) + PrefixLength : AlarmTagStripped.rindex(Suffix)]
				Success = AlarmKindMarker in Residue
			if Success:
				ResiduePrefix = Residue[:Residue.index(AlarmKindMarker)]
				ResidueSuffix = Residue[Residue.index(AlarmKindMarker) + len(AlarmKindMarker):]
			return Success, PrefixEndMarker, ResiduePrefix, ResidueSuffix, SuffixStartMarker

		def ImportAttribsIntoAlarm(ThisRow, ThisAlarm, ColAttribNames, SetpointColIndex,
			OverridePriorities=False):
			# import all available attribs from ThisRow (spreadsheet row)
			# into ThisAlarm (AlarmObjectInCore instance)
			# ColAttribNames (list of str): name of attrib to import into each column, or '' if no import into this column
			# SetpointColIndex (int): index of setpoint to import in row
			# OverridePriorities (bool): whether to set imported priorities as overrides
			# return UndoInfo (dict of AttribName: OldValue)
			assert isinstance(SetpointColIndex, int)
			UndoInfo = {}
			for ThisAttrib in AlarmAttrib.AllAlarmAttribs: # look for all possible attribs
				ThisAttribName = ThisAttrib.AttribName
				if ThisAttribName in ColAttribNames:
					ImportValue = ThisRow[ColAttribNames.index(ThisAttribName)] # incoming raw value from file
					UndoInfo[ThisAttribName] = getattr(ThisAlarm, ThisAttribName) # store old value for undo
					ThisAttribType = ThisAttrib.AttribType
					# import all directly importable attribs, one type at a time
					if ThisAttribType is str:
						setattr(ThisAlarm, ThisAttribName, ImportValue.strip())
					elif ThisAttribType is AlarmKindItem:
						setattr(ThisAlarm, ThisAttribName, GetAlarmKind(utilities.StripSpaces(ImportValue,
							CharsToStrip=' ' + chr(27)).lower(), ImportableOnly=True))
					elif ThisAttribType is bool:
						setattr(ThisAlarm, ThisAttribName, utilities.Bool2Str(utilities.StripSpaces(ImportValue,
							CharsToStrip=' ' + chr(27))))
					elif ThisAttribType is list: # expecting sequence of strings delimited by newlines. No mid-string stripping performed
						setattr(ThisAlarm, ThisAttribName, [a.strip() for a in ImportValue.split()])
					elif ThisAttribType is float:
						AttribToSet = getattr(ThisAlarm, ThisAttribName)
						# try to retrieve numerical value from import file
						ValueToSet = utilities.str2real(utilities.StripSpaces(ImportValue,
							CharsToStrip=' ' + chr(27)), meaninglessvalue=None)
						if ValueToSet is not None:
							AttribToSet.SetMyValue(ValueToSet)
							AttribToSet.SetMyUnit(ThisAttrib.Unit) # set expected unit for this attrib
					# import special attrib cases
			# import setpoint
			ImportedSetpoint = utilities.str2real(utilities.StripSpaces(ThisRow[SetpointColIndex],
				CharsToStrip=' ' + chr(27)), meaninglessvalue=None)
			if ImportedSetpoint is not None:
				UndoInfo[SetpointAttrib.AttribName] = copy.copy(ThisAlarm.Setpoint)
				ThisAlarm.Setpoint = ImportedSetpoint
			# import available response times. Find out whether we are using category or absolute value
			if ResponseTimeAvailableCategAttrib.AttribName in ColAttribNames: # using category
				UndoInfo[ResponseTimeAvailableCategAttrib] = ThisAlarm.ResponseTimeAvail.Value
				ThisAlarm.ResponseTimeAvail.Value = GetCategoryFromStr(
					Instr=ThisRow[ColAttribNames.index(ResponseTimeAvailableCategAttrib.AttribName)].strip().lower(),
					Categories=self.Urgencies, MatchAttribName='HumanName')
			else: # using absolute value
				ImportedValue = utilities.str2real(utilities.StripSpaces(ThisRow[ColAttribNames.index(ResponseTimeAvailableInMinAttrib.AttribName)],
					CharsToStrip=' ' + chr(27)), meaninglessvalue=None)
				if ImportedValue is not None:
					UndoInfo[ResponseTimeAvailableInMinAttrib] = copy.copy(ThisAlarm.ResponseTimeAvail)
					AttribToSet = getattr(ThisAlarm, ResponseTimeAvailableInMinAttrib.AttribName)
					AttribToSet.SetMyValue(ImportedValue)
					AttribToSet.SetMyUnit(ResponseTimeAvailableInMinAttrib.Unit) # set expected unit for this attrib
			# do same for required response times
			if ResponseTimeRequiredCategAttrib.AttribName in ColAttribNames: # using category
				UndoInfo[ResponseTimeRequiredCategAttrib] = ThisAlarm.ResponseTimeRequired.Value
				ThisAlarm.ResponseTimeRequired.Value = GetCategoryFromStr(
					Instr=ThisRow[ColAttribNames.index(ResponseTimeRequiredCategAttrib.AttribName)].strip().lower(),
					Categories=self.Urgencies, MatchAttribName='HumanName')
			else: # using absolute value
				ImportedValue = utilities.str2real(utilities.StripSpaces(ThisRow[ColAttribNames.index(ResponseTimeRequiredInMinAttrib.AttribName)],
					CharsToStrip=' ' + chr(27)), meaninglessvalue=None)
				if ImportedValue is not None:
					UndoInfo[ResponseTimeRequiredInMinAttrib] = copy.copy(ThisAlarm.ResponseTimeRequired)
					AttribToSet = getattr(ThisAlarm, ResponseTimeRequiredInMinAttrib.AttribName)
					AttribToSet.SetMyValue(ImportedValue)
					AttribToSet.SetMyUnit(ResponseTimeRequiredInMinAttrib.Unit) # set expected unit for this attrib
			# import severity
			if SeverityAttrib.AttribName in ColAttribNames:
				CandidateSeverities = [GetCategoryFromStr(Instr=a.strip().lower(), Categories=self.Severities,
					MatchAttribName='XMLName') for a in ThisRow[ColAttribNames.index(SeverityAttrib.AttribName)].split()]
				# did we recognise all severity categories imported?
				if CandidateSeverities.count(None) == 0:
					UndoInfo[SeverityAttrib] = copy.copy(ThisAlarm.Severities)
					# make a list of imported severities, same length as Causes
					ThisAlarm.Severities = utilities.PadList(CandidateSeverities, FieldWidth=len(ThisAlarm.Causes),
						PadValue=[s for s in self.Severities if getattr(s, 'IsUndefined', False)][0], Truncate=True)
			# import overridden priority
			if (PriorityAttrib.AttribName in ColAttribNames) and OverridePriorities:
				CandidatePriority = GetCategoryFromStr(Instr=ThisRow[ColAttribNames.index(PriorityAttrib.AttribName)].strip().lower(),
					Categories=self.Priorities, MatchAttribName='XMLName')
				if CandidatePriority is not None:
					UndoInfo[PriorityAttrib] = ThisAlarm.UserPriority.Value
					ThisAlarm.UserPriority.SetMyValue(CandidatePriority)
					# set "override" checkbox
					UndoInfo['OverridePriority'] = ThisAlarm.OverridePriority
					ThisAlarm.OverridePriority = True
			# import classifications
			ClassificationsChanged = False # whether to store undo record
			OldClassifications = ThisAlarm.Classifications
			for ThisClassif in self.Classifications:
				if ThisClassif.InternalName in ColAttribNames: # does the user want to import this classification?
					ClassificationsChanged = True
					ThisAlarm.Classifications[ThisClassif] = utilities.Bool2Str(ThisRow[ColAttribNames.index(ThisClassif.HumanName)])
			if ClassificationsChanged: UndoInfo['Classifications'] = OldClassifications
			return UndoInfo

		# main procedure for DoImport()
		ReturnInfo = {'ProblemReport': ''}
		# First, open file. data_only flag tries to load values rather than formulae. read_only reduces memory usage
		try:
			MyWorkbook = openpyxl.load_workbook(filename=Pathname, data_only=False, read_only=True)
		except:
			ReturnInfo['ProblemReport'] = _("Vizop couldn't read valid spreadsheet from file")
			return ReturnInfo
		# select required worksheet
		try:
			MyWorksheet = MyWorkbook['WorksheetName']
		except:
			ReturnInfo['ProblemReport'] = _("Vizop couldn't find specified worksheet in file")
			return ReturnInfo
		# ensure worksheet's max_row and max_column are set (some spreadsheet-making apps don't store them in the file)
		try:
			MyWorksheet.calculate_dimension(force=False)
		except ValueError: # max_row and max_column undefined; set them
			MyWorksheet.calculate_dimension(force=True) # could be time consuming for large files; do only if needed
		# find first row and column to read, and last valid column
		FirstRowIndex = MyWorksheet.min_row + HeaderRowCount
		FirstColIndex = MyWorksheet.min_column
		ValidColumnCount = MyWorksheet.max_column - MyWorksheet.min_column + 1
		# decide whether to deduce alarm tags, or read from file
		if (AlarmTagAttrib.AttribName in ColAttribNames) and not MultiLevel:
			DeduceAlarmTags = False
			AlarmTagColIndex = ColAttribNames.index(AlarmTagAttrib.AttribName) + FirstColIndex
		else: # check if a valid alarm tag prefix and suffix can be found
			DeduceAlarmTags = True
			Success, PrefixEndMarker, ResiduePrefix, ResidueSuffix, SuffixStartMarker =\
				GetStringsToDeduceAlarmTag(InstrumentTagTemplate, AlarmTagTemplate, AlarmKindMarker='H')
			if not Success:
				ReturnInfo['ProblemReport'] = _("Vizop can't use tag templates supplied")
				return ReturnInfo
		# if reading multilevel alarms, work out which levels to try to import for each row
		if MultiLevel:
			LevelsToTryToImport = [ThisAttrib for ThisAttrib in MultiLevelSetpointAttribs
				if ThisAttrib.AttribName in ColAttribNames]
			# make a list of which columns to try to import alarm level setpoints from
			LevelCols = [ColAttribNames.index(a.AttribName) for a in LevelsToTryToImport]
		else: # find alarm kind column
			AlarmKindColIndex = ColAttribNames.index(AlarmKindAttrib.AttribName)
		# make list of existing alarm tags in the alarm list
		ExistingAlarmTags = [a.AlarmTag for a in self.Alarms]
		# make counters for alarm import statistics
		RowsSuccessful = 0 # how many rows read successfully (even if all alarms in the row skipped due to not overwriting existing alarms)
		RowsFailed = 0 # how many rows couldn't be read due to lack of needed data such as alarm tag
		AlarmsCreated = 0 # how many alarms created and imported
		AlarmsOverwritten = 0 # how many existing alarms overwritten
		AlarmsNotOverwritten = 0 # how many alarms skipped because Overwrite is False
		UndoInfo = {} # dict with keys = AlarmTag, values = dict {AttribName: OldValue} for every data change
		# read rows
		for ThisRow in MyWorksheet.rows[FirstRowIndex:]:
			RowSuccess = True # whether we can import from this row
			# find which alarm levels to import from this row. We make KindsToImport, a list of tuples:
			# 	(alarm kind to import, column index to get setpoint from)
			if MultiLevel: # look for setpoints in columns labelled as SetpointHH/H/L/LL
				KindsToImport = [ (LevelsToTryToImport[i].AlarmKind, LevelCols[i]) for i in range(len(LevelsToTryToImport))
					if str(ThisRow[LevelCols[i]]).strip != '']
			else:
				# find which kind of alarm is in this row
				KindToImport = GetAlarmKind(ThisRow[AlarmKindColIndex])
				# if recognised, add to the import list
				if KindToImport: KindsToImport = [ (KindToImport, ColAttribNames.index(SetpointAttrib.AttribName)) ]
				else:
					KindsToImport = [] # can't import any alarm from this row
					RowSuccess = False
			if RowSuccess: # any alarms to import from this row? Get or make alarm tags
				AlarmTags = []
				if DeduceAlarmTags: # build alarm tags for each KindToImport
					# first, get the prefix from the instrument tag
					ThisRowInstrTag = ThisRow[ColAttribNames.index(InstrTagAttrib.AttribName)].strip()
					if PrefixEndMarker in ThisRowInstrTag:
						PrefixEndIndex = ThisRowInstrTag.index(PrefixEndMarker)
						ThisRowPrefix = ThisRowInstrTag[:PrefixEndMarker]
					else: # if PrefixEndMarker not found, can't import this row
						RowSuccess = False
					# get the suffix from the instrument tag
					if bool(SuffixStartMarker) and (SuffixStartMarker in ThisRowInstrTag[PrefixEndMarker:]):
						SuffixStartIndex = ThisRowInstrTag.rindex(SuffixStartMarker)
						ThisRowSuffix = ThisRowInstrTag[SuffixStartMarker:]
					else: ThisRowSuffix = ''
					if RowSuccess:
						for ThisKind, SetpointColIndex in KindsToImport:
							AlarmTags.append(ThisRowPrefix + ResiduePrefix + ThisKind.Tag + ResidueSuffix + ThisRowSuffix)
				else: # get alarm tag directly from row
					AlarmTags = [ThisRow[AlarmTagColIndex].strip()]
			if RowSuccess:
				RowsSuccessful += 1
				# work out where to actually import alarm(s) from this row, depending on whether alarm tags already exist in alarm list
				DataDestination = [] # alarm in self.Alarms to import incoming data to
				for ThisTag in AlarmTags:
					if ThisTag in ExistingAlarmTags: # alarm exists in alarm list
						if Overwrite:
							DataDestination.append(ThisTag)
							AlarmsOverwritten += 1
							UndoInfo[ThisTag] = {}
						else:
							DataDestination.append(None)
							AlarmsNotOverwritten += 1
					else: # alarm doesn't exist in alarm list. Create a new alarm
						DataDestination.append(CreateNewAlarm(HostingARObject=self, AlarmTag=ThisTag))
						AlarmsCreated += 1
						UndoInfo[ThisTag] = {}
				# import all data to alarms for this row, and capture changes in UndoInfo for this alarm tag
				for ThisKindIndex in range(len(AlarmTags)):
					if DataDestination[ThisKindIndex]: # import this kind of alarm? (it's a new alarm or overwritten existing alarm)
						UndoInfo[AlarmTags[ThisKindIndex]].update(ImportAttribsIntoAlarm(ThisRow,
							DataDestination[ThisKindIndex], ColAttribNames,
							SetpointColIndex=KindsToImport[ThisKindIndex][1]))
		return ReturnInfo

	def OpenAlarmImportFileTrial(self, Pathname=''):
		# open file PathName (confirmed as a readable .xlsx file) and get info from it to guide user in setting up an import
		# First, open file. data_only flag tries to load values rather than formulae. read_only reduces memory usage
		MyWorkbook = openpyxl.load_workbook(filename=Pathname, data_only=False, read_only=True)
		# get list of worksheets in the workbook
		WorksheetNames = MyWorkbook.sheetnames

class AlarmListHeader(object): pass

class AlarmList(display_utilities.ViewportBaseClass): # object containing all data needed to display alarm list
	# Each separate sub-object (header, cause etc) has attributes whose names are assumed to be same as in the data message from DataCore
	# NB this class has a forward definition earlier in this module.
	IsBaseClass = False # should be done for every subclass of ViewportBaseClass
	CanBeCreatedManually = True # whether the user should be able to create a Viewport of this class from scratch
	InternalName = 'AlarmList' # unique per class, used in messaging
	HumanName = _('Alarm list')
	PreferredKbdShortcut = 'L'
#	NewPHAObjRequired = ARObjectInCore  # which datacore PHA object class this Viewport spawns on creation.
#	# Should be None if the model shouldn't create a PHA object
	ImageSizeNoZoom = (20, 20)  # initial no-zoom size of all button images
	InitialEditPanelMode = 'Select'

	def __init__(self, **Args): # Alarm list initiation. Args must include Proj and can include DisplDevice and ParentWindow
		display_utilities.ViewportBaseClass.__init__(self, **Args)
		self.Proj = Args['Proj']
		self.PHAObj = None # instance of ARObjectInCore shown in this Viewport; set in DoNewViewport()
		self.DisplDevice = Args.get('DisplDevice', None)
#		self.Header = AlarmListHeader(MyAlarmList=self)
		self.Zoom = 1.0  # ratio of canvas coords to screen coords (absolute ratio, not %)
		self.PanX = self.PanY = 0  # offset of drawing origin, in screen coords
		self.OffsetX = self.OffsetY = 0  # offset of Viewport in display panel, in screen coords;
		# referenced in utilities.CanvasCoordsViewport() but not currently used
		self.ParentWindow = Args.get('DisplDevice', None)
#		self.BaseLayerBitmap = None  # a wx.Bitmap for holding the base layer of the FT, apart from floating objects such as the zoom widget
#		self.FloatingLayers = []  # a list of FloatLayer objects, for overlay onto the bitmap in self.Buffer, in arbitrary order
#		# set up images used in FT
#		self.ImagesNoZoom = {}
#		self.ArtProvider = art.ArtProvider()  # initialise art provider object
#		for ThisImageName in self.ArtProvider.ImageCatalogue(OnlyWithPrefix=info.FTImagePrefix):
#			# get the bitmap of each image, using image name as key including the prefix
#			self.ImagesNoZoom[ThisImageName] = self.ArtProvider.get_image(name=ThisImageName,
#																		  size=FTForDisplay.ImageSizeNoZoom,
#																		  conserve_aspect_ratio=True)
		# initialize zoom widget
		self.MyZoomWidget = display_utilities.ZoomWidgetObj(Viewport=self, InitialZoom=self.Zoom)
		self.LastElLClicked = self.LastElMClicked = self.LastElRClicked = None
		self.CurrentEditElement = None # which text component is currently being edited with a TextCtrl
		self.PaintNeeded = True # whether to execute DoRedraw() in display device's OnPaint() handler (bool)
		self.Panning = self.Zooming = False # whether user is currently changing display zoom or pan, for redraw efficiency

	def Wipe(self): # wipe all data in the alarm list and re-initialize
#		self.Header.InitializeData()
		pass

	def SetupDisplay(self, DisplDevice=None, ColScheme=None):
		# create grid and widgets for alarm list display in DisplDevice
		# ColScheme: vizop.ColourSchemeItem instance

		def SetupAlarmGrid(): pass
			# set up grid widget for display of alarm data. Not used yet

		# main procedure for SetupDisplay()
		assert isinstance(DisplDevice, wx.Panel)
		assert isinstance(ColScheme, display_utilities.ColourSchemeItem)
		self.DisplDevice = DisplDevice
		self.ColScheme = ColScheme
		self.DisplDevice.SetBackgroundColour(self.ColScheme.BackBright)
		MyControlFrame = self.DisplDevice.TopLevelFrame # make reference to overall frame
#		SettingsManager = settings.SettingsManager()
		self.StandardImageButtonSize = wx.Size(32, 32)
		NumFixedAlarmCols = 1 # how many alarm attribute columns are always visible
		NumVariableAlarmCols = 10 # how many variable alarm attribute columns are currently visible
		MaxNumVariableAlarmCols = 10 # how many variable alarm attribute columns can be visible

		# Make sizers to contain all widgets
		Sizer1 = wx.BoxSizer(orient=wx.VERTICAL) # overall sizer, contains sizers 2 and 3
		Sizer2 = wx.GridBagSizer(vgap=0, hgap=0) # control widgets
		Sizer3 = wx.GridSizer(vgap=0, hgap=0) # contains sizers 4, 5, 6 and alarm grid
		Sizer4 = wx.FlexGridSizer(rows=1, cols=4, vgap=0, hgap=0) # "select all" and scroll widgets
		Sizer5 = wx.FlexGridSizer(rows=1, cols=NumFixedAlarmCols + NumVariableAlarmCols, vgap=0, hgap=0) # filter widgets
		Sizer6 = wx.FlexGridSizer(rows=10, cols=2, vgap=0, hgap=0) # alarm selection widgets per alarm
		self.DisplDevice.SetSizer(Sizer1)
		Sizer1.Add(Sizer2)
		Sizer1.Add(Sizer3)

		# Make all the widgets for Sizer2
		# In Vizop, all widgets in sizers are in the form of a UIWidgetItem object
		self.MainHeaderLabel = UIWidgetItem(wx.StaticText(self.DisplDevice, -1, ''),
				ColLoc=0, ColSpan=4)
		self.ShowLabel = UIWidgetItem(wx.StaticText(self.DisplDevice, -1, _('Show:')),
				ColLoc=0, NewRow=True)
		self.IdentityGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Identity'), name='Identity'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=1, ColSpan=2)
		self.SetpointGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Setpoint'), name='Setpoint'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=3, ColSpan=1)
		self.OpInfoGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Operator information'), name='OpInfo'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=4, ColSpan=4)
		self.PriorityGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Priority'), name='Priority'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=1, ColSpan=2, NewRow=True)
		self.ClassifGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Classification'), name='Classif'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=3, ColSpan=1)
		self.TuningGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Tuning'), name='Tuning'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=4, ColSpan=2)
		self.AdvancedGroupCheck = UIWidgetItem(wx.CheckBox(self.DisplDevice, -1, _('Advanced'), name='Advanced'),
			Handler=self.OnAttribGroupCheck, Events=[wx.EVT_CHECKBOX], ColLoc=6, ColSpan=2)
		self.FiltersLabel = UIWidgetItem(wx.StaticText(self.DisplDevice, -1, _('Filters')),
				ColLoc=2, NewRow=True)
		# populate Sizer2
		Sizer2Widgets = [self.MainHeaderLabel,
			self.ShowLabel, self.IdentityGroupCheck, self.SetpointGroupCheck, self.OpInfoGroupCheck,
			self.PriorityGroupCheck, self.ClassifGroupCheck, self.TuningGroupCheck, self.AdvancedGroupCheck,
			self.FiltersLabel]
		display_utilities.PopulateSizer(Sizer=Sizer2, Widgets=Sizer2Widgets, ActiveWidgetList=self.WidgActive,
			DefaultFont=MyControlFrame.Fonts['NormalWidgetFont'],
			HighlightBkgColour=self.ColScheme.BackHighlight) # TODO: put text widgets into self.TextWidgActive

		# populate Sizer3
		Sizer3.Add(Sizer4)
		Sizer3.Add(Sizer5)
		Sizer3.Add(Sizer6)
		self.AlarmListGrid = DraggableGrid(Parent=self.DisplDevice)
		Sizer3.Add(self.AlarmListGrid)

		# populate Sizer4 with "select all" and scroll widgets
		self.SelectAllCheck = wx.CheckBox(self.DisplDevice, -1, '', name='SelectAll')
		self.DisplDevice.Bind(wx.EVT_CHECKBOX, self.OnSelectAllCheck)
		self.ScrollLeftButton = wx.Button(self.DisplDevice, size=self.StandardImageButtonSize, name='ScrollLeft')
		# TODO assign KeyStroke=[wx.WXK_CONTROL, wx.WXK_LEFT/RIGHT]
		self.ScrollLeftButton.Bind(wx.EVT_BUTTON, self.OnScrollHorizButton)
		self.ScrollLeftButton.SetBitmap(MyControlFrame.ButtonBitmap(wx.ART_GO_BACK))
		self.ScrollRightButton = wx.Button(self.DisplDevice, size=self.StandardImageButtonSize, name='ScrollRight')
		self.ScrollRightButton.Bind(wx.EVT_BUTTON, self.OnScrollHorizButton)
		self.ScrollRightButton.SetBitmap(MyControlFrame.ButtonBitmap(wx.ART_GO_FORWARD))
		Sizer4.AddMany( [self.SelectAllCheck, self.ScrollLeftButton, self.ScrollRightButton] )
#		self.AlarmTagFilterText = UIWidgetItem(wx.TextCtrl(self, -1), Handler=self.OnFilterText,
#			Events=[wx.EVT_TEXT], ColLoc=2, DataAttrib='AlarmTag')
#		# other filter widgets are created dynamically in BuildFilterWidgets()


	def OnAttribGroupCheck(self, Event): pass
	def OnSelectAllCheck(self, Event): pass
	def OnScrollHorizButton(self, Event): pass
	def BuildFilterWidgets(self): # make filter widgets for currently visible alarm attribute columns
		pass

	def StoreAllDataInXML(self, StartTag):
		# create an XML element as a subelement of StartTag (ElementTree.Element) and populate it with all Viewport
		# data required to be stored in project file.
		assert isinstance(StartTag, ElementTree.Element)
		# First, make top level XML element, and add common tags
		TopTag = projects.StoreViewportCommonDataInXML(Viewport=self, StartTag=StartTag)
		# TODO add any more tags that should be stored

ARObjectInCore.DefaultViewportType = AlarmList # we put this here because class ARObjectInCore is defined before class AlarmList
