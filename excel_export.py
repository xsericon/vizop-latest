# -*- coding: utf-8 -*-
# Module excel_export: part of Vizop, (c) 2020 xSeriCon
# contains common code for producing exported Excel files
import wx
from openpyxl import Workbook, worksheet
from openpyxl.worksheet import dimensions as Dimensions
from openpyxl.styles import PatternFill, Border, Side, alignment, Protection, Font, fills
import utilities, core_classes, info

DefaultFontColourRGB = (0,0,0)
MagicWidthRatio = 0.479 # mysterious ratio converting Excel column widths to actual values in mm

class ExcelTable_Table(object):
	# defines a block of cells in an Excel export. Each cell is one "Component" in the table. It's not necessary to
	# define a Component for every cell in the Table; any missing edge cells will be created (for drawing borders).

	def __init__(self, PositionRelativeTo=None, RelPosDirection='Right', GapToRight=0,
		SkipGapToRightIfAtRightOfSheet=True, GapBelow=0, SkipGapBelowIfAtBottomOfSheet=True, TopBorder=None,
		BottomBorder=None, LeftBorder=None, RightBorder=None):
		# PositionRelativeTo: another Table or None - the table this one should be oriented relative to
		# RelPosDirection (str): which direction this table should be oriented relative to the one named above
		# GapToRight (int/float): if nonzero, one blank cell with this relative width is left blank to the right of this table
		# SkipGapToRightIfAtRightOfSheet (bool): don't leave gap if there's no table to the right
		# GapBelow (int): if nonzero, one blank cell with height = this number of lines is left blank below this table
		# SkipGapBelowIfAtBottomOfSheet (bool): don't leave gap if there's no table below
		# Borders: border object or None; borders to be drawn all round the table
		assert isinstance(PositionRelativeTo, ExcelTable_Table) or (PositionRelativeTo is None)
		assert RelPosDirection in [info.RightLabel, info.BelowLabel]
		assert isinstance(GapToRight, (int, float))
		assert GapToRight >= 0
		assert isinstance(SkipGapToRightIfAtRightOfSheet, bool)
		assert isinstance(GapBelow, int)
		assert GapBelow >= 0
		assert isinstance(SkipGapBelowIfAtBottomOfSheet, bool)
		assert isinstance(TopBorder, ExcelTable_Border) or (TopBorder is None)
		assert isinstance(BottomBorder, ExcelTable_Border) or (BottomBorder is None)
		assert isinstance(LeftBorder, ExcelTable_Border) or (LeftBorder is None)
		assert isinstance(RightBorder, ExcelTable_Border) or (RightBorder is None)
		self.PositionRelativeTo = PositionRelativeTo
		self.RelPosDirection = RelPosDirection
		self.GapToRight = GapToRight
		self.SkipGapToRightIfAtRightOfSheet = SkipGapToRightIfAtRightOfSheet
		self.GapBelow = GapBelow
		self.SkipGapBelowIfAtBottomOfSheet = SkipGapBelowIfAtBottomOfSheet
		self.TopBorder = TopBorder
		self.BottomBorder = BottomBorder
		self.LeftBorder = LeftBorder
		self.RightBorder = RightBorder
		self.Components = [] # list of ExcelTable_Component instances

	def PopulateTable(self, WS, StartRow, StartCol, RelativeWidths):
		# populate cells in worksheet WS based on self.Components, starting from StartRow and StartCol (2 x int, 1-based)
		# RelativeWidths (list of int/float): relative widths assigned so far to all columns in WS.
		# This method does not apply any borders; that's done in self.DrawBorders()
		# return new RelativeWidths, next available Row, next available Col
		assert isinstance(WS, worksheet.worksheet.Worksheet)
		assert isinstance(StartRow, int)
		assert StartRow > 0
		assert isinstance(StartCol, int)
		assert StartCol > 0
		assert isinstance(RelativeWidths, list)
		self.StartRow = StartRow # needed for any table whose position is relative to this one
		self.StartCol = StartCol
		# start with width values supplied; will be overwritten by any larger values in this Table
		MaxRelWidthsRequested = RelativeWidths[:]
		MergeGroups = [] # each item is a list of Components whose cells are to be merged together
		MergeCoords = {} # keys: (Row, Col) coord of a cell that's part of a MergeGroup;
			# values: index in MergeGroups containing the Component from which this
			# cell is merged (i.e. the one above or to the left)
		for ThisComponent in self.Components:
			# find which component to use as a positional reference
			if ThisComponent.PositionRelativeTo:
				RefCpt = ThisComponent.PositionRelativeTo
				# confirm reference component has already been processed
				assert (RefCpt.Row is not None) and (RefCpt.Col is not None)
				if ThisComponent.RelPosDirection == info.RightLabel:
					ThisComponent.Row = ThisComponent.PositionRelativeTo.Row
					ThisComponent.Col = ThisComponent.PositionRelativeTo.Col + 1
				elif ThisComponent.RelPosDirection == info.BelowLabel:
					ThisComponent.Row = ThisComponent.PositionRelativeTo.Row + 1
					ThisComponent.Col = ThisComponent.PositionRelativeTo.Col
			else: # if no reference component specified, put this component at the top left of the table
				ThisComponent.Row = StartRow
				ThisComponent.Col = StartCol
			# make a cell to contain the component's content
			ThisCell = WS.cell(row=ThisComponent.Row, column=ThisComponent.Col, value=ThisComponent.Content)
			# apply cell background colour
			ThisCell.fill = PatternFill('solid', fgColor=utilities.HexTuple2str(ThisComponent.BackgColour))
			# apply alignment
			ThisCell.alignment = alignment.Alignment(horizontal={info.LeftLabel: 'left', info.CentreLabel: 'center',
				info.RightLabel: 'right'}[ThisComponent.HorizAlignment],
				vertical={info.TopLabel: 'top', info.CentreLabel: 'center',
				info.BottomLabel: 'bottom'}[ThisComponent.VertAlignment], wrapText=ThisComponent.Wrap)
			# apply font
			if ThisComponent.FontStyle:
				ThisCell.font = Font(name=ThisComponent.FontStyle.GetFaceName(),
					size=ThisComponent.FontStyle.GetPointSize(),
					bold=ThisComponent.FontStyle.GetWeight() == wx.FONTWEIGHT_BOLD,
					italic=ThisComponent.FontStyle.GetStyle() == wx.FONTSTYLE_ITALIC,
					vertAlign=None, underline='none', strike=False,
					color='FF' + utilities.HexTuple2str(getattr(ThisComponent.FontStyle, 'Colour', DefaultFontColourRGB)))
			# apply merging. First, find whether this cell's coordinates are already in MergeCoords
			ThisMergeGroupIndex = None # index in MergeGroups to which this Component belongs
			if (ThisComponent.Row, ThisComponent.Col) in MergeCoords.keys():
				# this cell is already marked for merging. Add it to a merge group with the cell from which it is merged
				ThisMergeGroupIndex = MergeCoords[(ThisComponent.Row, ThisComponent.Col)]
				MergeGroups[ThisMergeGroupIndex].append(ThisComponent)
			# Check if this component is requesting merge with another component
			if ThisComponent.MergeToRight:
				# if this cell isn't already in a merge group, create a new merge group and add this cell to it
				if ThisMergeGroupIndex is None:
					MergeGroups.append([ThisComponent])
					ThisMergeGroupIndex = len(MergeGroups) - 1
				# add the cell to the right to MergeCoords, so that it will be added to the merge group when it's processed
				MergeCoords[(ThisComponent.Row, ThisComponent.Col + 1)] = ThisMergeGroupIndex
			if ThisComponent.MergeDown:
				# if this cell isn't already in a merge group, create a new merge group and add this cell to it
				if ThisMergeGroupIndex is None:
					MergeGroups.append([ThisComponent])
					ThisMergeGroupIndex = len(MergeGroups) - 1
				# add the cell below to MergeCoords, so that it will be added to the merge group when it's processed
				MergeCoords[(ThisComponent.Row + 1, ThisComponent.Col)] = ThisMergeGroupIndex
			# store the requested width for this cell
			# - remembering that Col is 1-based, so col 1 width is at MaxRelWidthsRequested[0]
			# first, make sure there's enough column entries in MaxRelWidthsRequested
			MaxRelWidthsRequested.extend([0.0] * (ThisComponent.Col - len(MaxRelWidthsRequested)))
			MaxRelWidthsRequested[ThisComponent.Col - 1] = \
				max(MaxRelWidthsRequested[ThisComponent.Col - 1], ThisComponent.RelWidth)
		# do cell merge of cells in MergeGroups
		for ThisMG in MergeGroups:
			if len(ThisMG) > 1: # proceed only if the merge group contains more than one cell
				# find first and last row and column in the merge group
				StartRow = min(c.Row for c in ThisMG)
				EndRow = max(c.Row for c in ThisMG)
				StartCol = min(c.Col for c in ThisMG)
				EndCol = max(c.Col for c in ThisMG)
				WS.merge_cells(start_row=StartRow, start_column=StartCol, end_row=EndRow, end_column=EndCol)
		# find last, and next available, row and column
		self.EndRow = max(c.Row for c in self.Components)
		self.EndCol = max(c.Col for c in self.Components)
		if self.GapToRight:
			# allow space for one more column to the right, with relative width at least GapToRight
			self.NextTableStartCol = self.EndCol + 2
			MaxRelWidthsRequested.extend([0.0] * (1 + self.EndCol - len(MaxRelWidthsRequested)))
			MaxRelWidthsRequested[self.EndCol + 1] = max(MaxRelWidthsRequested[self.EndCol + 1], self.GapToRight)
		else: self.NextTableStartCol = self.EndCol + 1
		self.NextTableStartRow = self.EndRow + 2 if self.GapBelow else self.EndRow + 1
		return MaxRelWidthsRequested

	def DrawBorders(self, WS):
		# draw borders around table, and around components within table.
		# Does not draw borders in inter-table gaps; that's done in Sheet.SetTableBorders()
		# First, make lists of row coords in left and right table edge cells
		LeftEdgeRows = [r for r in range(self.StartRow, self.EndRow + 1)]
		RightEdgeRows = LeftEdgeRows[:]
		# likewise for top and bottom edge cells, except corners as they're already covered by left and right edges
		TopEdgeCols = [c for c in range(self.StartCol + 1, self.EndCol)]
		BottomEdgeCols = TopEdgeCols[:]
		for ThisComponent in self.Components:
			# for each component, find out if it's at the top, bottom, left or right edge of the table.
			# If so, draw table border on that edge. Otherwise, draw component's own border, if any
			BorderArgs = {}
			if ThisComponent.Col == self.StartCol:
				BorderArgs['left'] = self.LeftBorder.Side
				# the "if" is to avoid crash in case there's 2 cells at the same location, oops
				if ThisComponent.Row in LeftEdgeRows: LeftEdgeRows.remove(ThisComponent.Row)
			elif ThisComponent.LeftBorder: BorderArgs['left'] = ThisComponent.LeftBorder.Side
			if ThisComponent.Col == self.EndCol:
				BorderArgs['right'] = self.RightBorder.Side
				if ThisComponent.Row in RightEdgeRows: RightEdgeRows.remove(ThisComponent.Row)
			elif ThisComponent.RightBorder: BorderArgs['right'] = ThisComponent.RightBorder.Side
			if ThisComponent.Row == self.StartRow:
				BorderArgs['top'] = self.TopBorder.Side
				if ThisComponent.Col in TopEdgeCols: TopEdgeCols.remove(ThisComponent.Col)
			elif ThisComponent.TopBorder: BorderArgs['top'] = ThisComponent.TopBorder.Side
			if ThisComponent.Row == self.EndRow:
				BorderArgs['bottom'] = self.BottomBorder.Side
				if ThisComponent.Col in BottomEdgeCols: BottomEdgeCols.remove(ThisComponent.Col)
			elif ThisComponent.BottomBorder: BorderArgs['bottom'] = ThisComponent.BottomBorder.Side
			# apply the borders, if any
			if BorderArgs:
				ThisCell = WS.cell(row=ThisComponent.Row, column=ThisComponent.Col)
				ThisCell.border = Border(**BorderArgs)
		# apply table border to any missing edge cells, in the order (left, right), (top, bottom)
		for (ThisRowList, ThisEdge, ThisBorder, ThisCol) in [(LeftEdgeRows, 'left', self.LeftBorder, self.StartCol),
				(RightEdgeRows, 'right', self.RightBorder, self.EndCol)]:
			for EdgeRow in ThisRowList:
				ThisCell = WS.cell(row=EdgeRow, column=ThisCol)
				BorderArgs = {ThisEdge: ThisBorder.Side}
				# add top or bottom border if this cell is at the top or bottom of the table
				if EdgeRow == self.StartRow: BorderArgs['top'] = self.TopBorder.Side
				if EdgeRow == self.EndRow: BorderArgs['bottom'] = self.BottomBorder.Side
				ThisCell.border = Border(**BorderArgs)
		for (ThisColList, ThisEdge, ThisBorder, ThisRow) in [(TopEdgeCols, 'top', self.TopBorder, self.StartRow),
				(BottomEdgeCols, 'bottom', self.BottomBorder, self.EndRow)]:
			for EdgeCol in ThisColList:
				ThisCell = WS.cell(row=ThisRow, column=EdgeCol)
				BorderArgs = {ThisEdge: ThisBorder.Side}
				ThisCell.border = Border(**BorderArgs)

class ExcelTable_Component(object):
	# defines a single cell in a table comprising part of an Excel export

	def __init__(self, PositionRelativeTo=None, RelPosDirection='Right', TopBorder=None,
		BottomBorder=None, LeftBorder=None, RightBorder=None, Content='', VertAlignment=info.TopLabel,
		HorizAlignment=info.LeftLabel, Wrap=True, LeftIndentInmm=0, LeftIndentInRelWidth=0, FontStyle=None,
		BackgColour=(255,255,255), RelWidth=1.0, MergeToRight=False, MergeDown=False):
		# PositionRelativeTo: another Component or None - the component this one should be oriented relative to
		# RelPosDirection (str): which direction this component should be oriented relative to the one named above
		# Borders: border object or None
		# Content: text to put in the cell
		# VertAlignment, HorizAlignment: text alignment in the cell
		# Wrap (bool): whether to apply word wrap to text
		# LeftIndent: text indent to apply. LeftIndentInmm takes priority.
		# FontStyle: a wx.Font object or None. The wx.Font object can optionally have attrib 'Colour' (rgb tuple)
		# RelWidth: relative width of cell relative to all other cells in the sheet
		# Merge: whether to merge with any adjacent cell in specified direction
		assert isinstance(PositionRelativeTo, ExcelTable_Component) or (PositionRelativeTo is None)
		assert RelPosDirection in [info.RightLabel, info.BelowLabel]
		assert isinstance(TopBorder, ExcelTable_Border) or (TopBorder is None)
		assert isinstance(BottomBorder, ExcelTable_Border) or (BottomBorder is None)
		assert isinstance(LeftBorder, ExcelTable_Border) or (LeftBorder is None)
		assert isinstance(RightBorder, ExcelTable_Border) or (RightBorder is None)
		assert isinstance(Content, str), 'Creating Component with Content %s, type %s' % (str(Content), str(type(Content)))
		assert VertAlignment in [info.TopLabel, info.CentreLabel, info.BottomLabel]
		assert HorizAlignment in [info.LeftLabel, info.CentreLabel, info.RightLabel]
		assert isinstance(Wrap, bool)
		assert isinstance(LeftIndentInmm, (int, float))
		assert LeftIndentInmm >= 0
		assert isinstance(LeftIndentInRelWidth, (int, float))
		assert LeftIndentInRelWidth >= 0
		assert isinstance(FontStyle, wx.Font) or (FontStyle is None)
		assert isinstance(BackgColour, tuple)
		assert isinstance(RelWidth, (int, float))
		assert RelWidth > 0
		assert isinstance(MergeToRight, bool)
		assert isinstance(MergeDown, bool)
		self.PositionRelativeTo = PositionRelativeTo
		self.RelPosDirection = RelPosDirection
		self.TopBorder = TopBorder
		self.BottomBorder = BottomBorder
		self.LeftBorder = LeftBorder
		self.RightBorder = RightBorder
		self.Content = Content
		self.VertAlignment = VertAlignment
		self.HorizAlignment = HorizAlignment
		self.Wrap = Wrap
		self.LeftIndentInmm = LeftIndentInmm
		self.LeftIndentInRelWidth = LeftIndentInRelWidth
		self.FontStyle = FontStyle
		self.BackgColour = BackgColour
		self.RelWidth = RelWidth
		self.MergeToRight = MergeToRight
		self.MergeDown = MergeDown
		self.Row = self.Col = None # actual row and col address in the worksheet. Set in PopulateTable()

class ExcelTable_Sheet(object):
	# defines a block of tables defining content of an exported spreadsheet

	def __init__(self, TabName, TabColour, TargetWidth):
		# TabName (str): Name to be shown on Excel worksheet tab. Can't be blank
		# TabColour (rgb tuple, 3 x int): Colour to be applied to Excel worksheet tab
		# TargetWidth (int/float): Total width of all cells in the Sheet, in mm. Gets rounded to the nearest mm here
		assert isinstance(TabName, str)
		assert TabName.strip() # ensure it's not blank or whitespace only
		assert isinstance(TabColour, tuple)
		assert isinstance(TargetWidth, (int, float))
		assert TargetWidth > 0
		object.__init__(self)
		self.TabName = TabName
		self.TabColour = TabColour
		self.TargetWidth = int(round(TargetWidth))
		self.Tables = [] # ExcelTable_Table instances

	def PopulateSheet(self, WS, AbsSheetWidth, StartRow=1, StartCol=1):
		# populate all Components into their respective Tables, then assemble Tables into worksheet WS
		# The first Table is placed at cell StartRow, StartCol (1-based)
		# Cell widths will be adjusted so that total sheet width is AbsSheetWidth in mm
		assert isinstance(WS, worksheet.worksheet.Worksheet)
		assert isinstance(AbsSheetWidth, (int, float))
		assert AbsSheetWidth > 0
		assert isinstance(StartRow, int)
		assert StartRow > 0
		assert isinstance(StartCol, int)
		assert StartCol > 0
		MaxRelWidthsSoFar = [] # column relative widths requested by cells
		# populate each table
		for ThisTable in self.Tables:
			# set position for table, based on PositionRelativeTo and RelPosDirection
			if ThisTable.PositionRelativeTo is None: # start at initial position in sheet
				ThisTableStartRow = StartRow
				ThisTableStartCol = StartCol
			else: # start at position relative to another table
				if ThisTable.RelPosDirection == info.RightLabel:
					# position this table on the same row, and to the right of the reference table
					ThisTableStartRow = ThisTable.PositionRelativeTo.StartRow
					ThisTableStartCol = ThisTable.PositionRelativeTo.NextTableStartCol
				elif ThisTable.RelPosDirection == info.BelowLabel:
					# position this table below, and starting at the same column as, the reference table
					ThisTableStartRow = ThisTable.PositionRelativeTo.NextTableStartRow
					ThisTableStartCol = ThisTable.PositionRelativeTo.StartCol
			MaxRelWidthsSoFar = ThisTable.PopulateTable(WS=WS,
				StartRow=ThisTableStartRow, StartCol=ThisTableStartCol, RelativeWidths=MaxRelWidthsSoFar)
		# set column widths. First, find conversion factor from relative widths to absolute widths.
		# In following row, min(1, x) is to avoid embarrassing divide-by-zero crash
		ColWidthRatio = MagicWidthRatio * AbsSheetWidth / max(1, sum(MaxRelWidthsSoFar))
		for (ThisCol, ThisWidth) in enumerate(MaxRelWidthsSoFar):
			# below, the call to HumanValue() gets a sequence of letters 'A', 'B' etc
			WS.column_dimensions[core_classes.UpperCaseLetterNumberSystem.HumanValue(TargetValue=ThisCol + 1)].width = \
				ThisWidth * ColWidthRatio
		# set borders around and between tables
		# TODO consider adding dummy cells (containing whitespace only) between tables, so that user can select the
		# entire Sheet in Excel using Ctrl + A
		self.SetTableBorders(WS=WS)

	def SetTableBorders(self, WS):
		# set borders around and between tables
		for ThisTable in self.Tables:
			ThisTable.DrawBorders(WS=WS)
		# still TODO: draw borders between tables

class ExcelTable_Border(object):
	# defines the appearance of a border along one edge of a cell

	def __init__(self, Style, Colour):
		# Style (str): appearance of border
		# Colour (rgb tuple)
		assert Style in ['hair', 'dashed', 'mediumDashDot', 'mediumDashDotDot', 'slantDashDot', 'double', 'thick',
			'mediumDashed', 'thin', 'medium', 'dashDotDot', 'dashDot', 'dotted']
		assert isinstance(Colour, tuple)
		self.Style = Style
		self.Colour = Colour
		# define an openpxyl Side instance to represent this border - defined once only, on init
		self.Side = Side(border_style=Style, color=utilities.HexTuple2str(Colour))

def SetupWorkbook():
	# create and return a workbook object containing one empty worksheet
	return Workbook()

def SetupWorksheet(WB, TabName, TabColour=(255,255,255)):
	# create a worksheet object at the end of workbook WB, set it as WB's active worksheet, and return the worksheet
	assert isinstance(WB, Workbook)
	assert isinstance(TabName, str)
	assert TabName.strip() # ensure it's not empty or whitespace only
	assert isinstance(TabColour, tuple)
	ThisWS = WB.create_sheet()
	WB.active = ThisWS
	ThisWS.title = TabName
	ThisWS.sheet_properties.tabColor = utilities.HexTuple2str(TabColour)
	return ThisWS

def WriteWBToFile(WB, FilePath):
	# write workbook WB to file in FilePath
	# FilePath is assumed to be a writeable path
	WB.save(filename=FilePath)

def MergeAndFillCells(Table, StartComponent, TotalCols, FinalRightBorder=None):
	# in Table, append after StartComponent a series of Components, on the right of StartComponent and merged to it
	# TotalCols (int): Total number of columns (including StartComponent's) to be merged into one cell
	# FinalRightBorder: right border to apply to the last cell added
	# return the final component added, if any. If no components added, return StartComponent. This may be needed as
	# a positional reference for further components in the table
	# Inserted cells have the same top and bottom border as StartComponent
	assert isinstance(Table, ExcelTable_Table)
	assert isinstance(StartComponent, ExcelTable_Component)
	assert StartComponent in Table.Components
	assert isinstance(TotalCols, int)
	assert TotalCols >= 1
	assert isinstance(FinalRightBorder, ExcelTable_Border) or (FinalRightBorder is None)
	# first, find index of StartComponent in Table; we don't assume it's at the end of Table.Components
	StartComponentIndex = Table.Components.index(StartComponent)
	# insert empty cells up to the final cell, if TotalCols > 1
	ComponentsAdded = False
	for InsertIndex in range(StartComponentIndex + 1, StartComponentIndex + TotalCols):
		IsFinalComponent = (InsertIndex == StartComponentIndex + TotalCols - 1)
		ComponentsAdded = True # whether any components added to Table
		Table.Components.insert(InsertIndex, ExcelTable_Component(MergeToRight=not IsFinalComponent,
			PositionRelativeTo=Table.Components[InsertIndex - 1], RelPosDirection=info.RightLabel,
			TopBorder=StartComponent.TopBorder, BottomBorder=StartComponent.BottomBorder,
			RightBorder = FinalRightBorder if IsFinalComponent else None))
		if IsFinalComponent: ComponentToReturn = Table.Components[InsertIndex] # store rightmost new component
	# if any cells were added, set StartComponent to merge to right
	if ComponentsAdded: StartComponent.MergeToRight = True
	# if no cells were added, apply FinalRightBorder to StartComponent
	else:
		StartComponent.RightBorder = FinalRightBorder
		ComponentToReturn = StartComponent
	return ComponentToReturn
